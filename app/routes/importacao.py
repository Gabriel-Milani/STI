from sqlite3 import IntegrityError

from flask import Blueprint, current_app, request
from openpyxl import load_workbook
from ..database import get_db
from ..services.auth_utils import login_required, current_user_id, current_user_name
from ..services.helpers import api_ok, api_error
from ..services.products import ProductCreateError, ProductValidationError, create_product, get_active_location, normalize_product_payload

importacao_bp = Blueprint("importacao", __name__)

EXPECTED_HEADERS = [
    "nome", "categoria", "modelo", "codigo_barras",
    "quantidade_inicial", "estoque_minimo", "localizacao_codigo", "observacao"
]
ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}


def allowed_excel_file(filename):
    name = (filename or "").lower()
    return any(name.endswith(ext) for ext in ALLOWED_EXTENSIONS)


@importacao_bp.get("/template")
@login_required
def template_info():
    return api_ok({"headers": EXPECTED_HEADERS})


@importacao_bp.post("/produtos")
@login_required
def importar_produtos():
    if "arquivo" not in request.files:
        return api_error("Envie um arquivo Excel no campo 'arquivo'.", 400)
    file = request.files["arquivo"]
    if not allowed_excel_file(file.filename):
        return api_error("Envie um arquivo Excel .xlsx ou .xlsm.", 400)
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception:
        return api_error("Arquivo Excel inválido.", 400)

    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    missing = [h for h in EXPECTED_HEADERS if h not in headers]
    if missing:
        return api_error("Cabeçalhos ausentes na planilha.", 400, missing)
    idx = {h: headers.index(h) for h in EXPECTED_HEADERS}

    erros = []
    rows_to_import = []
    seen_barcodes = {}
    with get_db() as db:
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = {h: row[idx[h]] for h in EXPECTED_HEADERS}
            nome = str(data.get("nome") or "").strip()
            if not nome:
                continue
            data["recebido_por"] = current_user_name()
            try:
                payload = normalize_product_payload(data)
                if not get_active_location(db, payload):
                    raise ProductValidationError("Escolha uma localização para o produto.")
                barcode = payload["codigo_barras"]
                if barcode:
                    previous_row = seen_barcodes.get(barcode)
                    if previous_row:
                        raise ProductValidationError(f"Código de barras repetido na linha {previous_row}.")
                    seen_barcodes[barcode] = row_num
                    exists = db.execute("SELECT id FROM produtos WHERE codigo_barras = ?", (barcode,)).fetchone()
                    if exists:
                        raise ProductValidationError("Código de barras já cadastrado.")
                rows_to_import.append(data)
            except ProductValidationError as error:
                erros.append({"linha": row_num, "erro": error.message})
        if erros:
            db.rollback()
            return api_error("Importação não realizada. Corrija os erros da planilha e tente novamente.", 400, erros)

        criados = 0
        try:
            for data in rows_to_import:
                create_product(
                    db,
                    data,
                    current_user_id(),
                    audit_action="importar",
                    initial_note="Importação inicial",
                )
                criados += 1
            db.commit()
        except ProductCreateError as error:
            db.rollback()
            return api_error("Importação não realizada.", 400, [{"erro": error.message}])
        except IntegrityError:
            db.rollback()
            return api_error("Importação não realizada. Verifique códigos duplicados.", 400)
        except Exception:
            current_app.logger.exception("Erro ao importar produtos")
            db.rollback()
            return api_error("Não foi possível importar os produtos.", 500)
    return api_ok({"criados": criados, "erros": []}, "Importação finalizada.")
